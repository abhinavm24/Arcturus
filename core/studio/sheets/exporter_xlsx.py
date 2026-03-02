"""XLSX export via openpyxl for sheet artifacts with visual polish and charts."""

from __future__ import annotations

import hashlib
import logging
import re
import statistics
from dataclasses import dataclass
from pathlib import Path
from typing import TYPE_CHECKING, Any, Dict, List, Optional

if TYPE_CHECKING:
    from openpyxl.worksheet.worksheet import Worksheet

from core.schemas.studio_schema import SheetAnalysisReport, SheetContentTree, SheetTab

logger = logging.getLogger(__name__)

# Characters invalid in Excel worksheet names
_INVALID_CHARS = re.compile(r"[\[\]:*?/\\]")
_MAX_SHEET_NAME_LEN = 31

_VALID_VISUAL_PROFILES = {"balanced", "conservative", "max"}
_PROFILE_MAX_CHARTS = {
    "conservative": 1,
    "balanced": 3,
    "max": 5,
}

_CURRENCY_KEYWORDS = {
    "amount",
    "revenue",
    "cost",
    "price",
    "budget",
    "profit",
    "expense",
    "sales",
}
_PERCENT_KEYWORDS = {"pct", "percent", "%", "growth", "rate", "ratio", "margin"}
_TOTAL_KEYWORDS = {"total", "subtotal", "grand total", "average", "avg"}
_MONTH_TOKENS = {
    "jan", "feb", "mar", "apr", "may", "jun", "jul", "aug", "sep", "oct", "nov", "dec",
    "january", "february", "march", "april", "june", "july", "august", "september", "october", "november", "december",
}


@dataclass(frozen=True)
class SheetPalette:
    id: str
    name: str
    primary: str
    secondary: str
    accent: str
    background: str
    text: str
    header_text: str
    zebra: str
    subtotal: str
    trend_up: str
    trend_down: str
    trend_flat: str
    header_contrast: str  # "light-on-dark" or "dark-on-light"


_PALETTES: List[SheetPalette] = [
    SheetPalette(
        id="slate-executive",
        name="Slate Executive",
        primary="2D3748",
        secondary="4A5568",
        accent="4A7AB5",
        background="F7FAFC",
        text="1A202C",
        header_text="FFFFFF",
        zebra="EDF2F7",
        subtotal="E2E8F0",
        trend_up="C6DAF0",
        trend_down="FED7D7",
        trend_flat="E2E8F0",
        header_contrast="light-on-dark",
    ),
    SheetPalette(
        id="iron-neutral",
        name="Iron Neutral",
        primary="1F2937",
        secondary="374151",
        accent="6B7280",
        background="F9FAFB",
        text="111827",
        header_text="FFFFFF",
        zebra="F3F4F6",
        subtotal="E5E7EB",
        trend_up="D1D5DB",
        trend_down="FEE2E2",
        trend_flat="E5E7EB",
        header_contrast="light-on-dark",
    ),
    SheetPalette(
        id="sand-warm",
        name="Sand Warm",
        primary="44403C",
        secondary="57534E",
        accent="B08D57",
        background="FAF9F6",
        text="292524",
        header_text="FFFFFF",
        zebra="F5F5F4",
        subtotal="E7E5E4",
        trend_up="E8DCCC",
        trend_down="FECACA",
        trend_flat="E7E5E4",
        header_contrast="light-on-dark",
    ),
]
_PALETTE_BY_ID = {p.id: p for p in _PALETTES}


def sanitize_sheet_name(name: str) -> str:
    """Sanitize a worksheet name for Excel compatibility."""
    cleaned = _INVALID_CHARS.sub("", name)
    return cleaned[:_MAX_SHEET_NAME_LEN] if cleaned else "Sheet"


def _is_numeric(value: Any) -> bool:
    return isinstance(value, (int, float)) and not isinstance(value, bool)


def _normalize_visual_profile(metadata: Optional[Dict[str, Any]]) -> str:
    if not isinstance(metadata, dict):
        return "balanced"
    profile = metadata.get("visual_profile")
    if not isinstance(profile, str):
        return "balanced"
    profile = profile.strip().lower()
    return profile if profile in _VALID_VISUAL_PROFILES else "balanced"


def _select_palette(content_tree: SheetContentTree) -> SheetPalette:
    metadata = content_tree.metadata if isinstance(content_tree.metadata, dict) else {}
    palette_hint = metadata.get("palette_hint") if metadata else None
    if isinstance(palette_hint, str):
        palette_hint = palette_hint.strip().lower()
        if palette_hint in _PALETTE_BY_ID:
            return _PALETTE_BY_ID[palette_hint]

    digest = hashlib.sha256(content_tree.workbook_title.encode("utf-8")).hexdigest()
    idx = int(digest[:8], 16) % len(_PALETTES)
    return _PALETTES[idx]


def _header_index(tab: SheetTab, header_name: str) -> Optional[int]:
    needle = header_name.strip().lower()
    for i, header in enumerate(tab.headers):
        if header.strip().lower() == needle:
            return i
    return None


def _numeric_columns(tab: SheetTab) -> List[int]:
    cols: List[int] = []
    for col_idx in range(len(tab.headers)):
        if any(col_idx < len(row) and _is_numeric(row[col_idx]) for row in tab.rows):
            cols.append(col_idx)
    return cols


def _categorical_columns(tab: SheetTab) -> List[int]:
    cols: List[int] = []
    for col_idx in range(len(tab.headers)):
        values = [row[col_idx] for row in tab.rows if col_idx < len(row)]
        string_values = [v for v in values if isinstance(v, str) and v.strip()]
        if string_values and len(string_values) >= max(2, len(values) // 2):
            cols.append(col_idx)
    return cols


def _looks_temporal(values: List[Any], header: str) -> bool:
    h = header.strip().lower()
    if "date" in h or "month" in h or "year" in h or "week" in h or "quarter" in h:
        return True
    tokens = {str(v).strip().lower() for v in values if isinstance(v, str) and v.strip()}
    return len(tokens.intersection(_MONTH_TOKENS)) > 0


def _preferred_tab_score(name: str) -> int:
    lower = name.lower()
    score = 0
    if any(k in lower for k in ("summary", "pivot", "stats", "kpi", "overview", "dashboard")):
        score += 10
    if any(k in lower for k in ("analysis", "report", "trend")):
        score += 5
    return score


def _infer_number_format(header: str, values: List[Any]) -> Optional[str]:
    numeric_vals = [v for v in values if _is_numeric(v)]
    if not numeric_vals:
        return None

    lower = header.strip().lower()
    if any(token in lower for token in _PERCENT_KEYWORDS):
        return "0.00%"
    if any(token in lower for token in _CURRENCY_KEYWORDS):
        return '"$"#,##0.00'
    if all(float(v).is_integer() for v in numeric_vals):
        return "#,##0"
    return "#,##0.00"


def _build_border():
    from openpyxl.styles import Border, Side

    thin = Side(style="thin", color="D1D5DB")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _apply_structural_formatting(ws: Worksheet, tab: SheetTab, palette: SheetPalette) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    if not tab.headers:
        return

    max_row = len(tab.rows) + 1
    max_col = len(tab.headers)
    max_col_letter = get_column_letter(max_col)

    header_fill = PatternFill(start_color=palette.primary, end_color=palette.primary, fill_type="solid")
    header_font = Font(bold=True, color=palette.header_text)
    zebra_fill = PatternFill(start_color=palette.zebra, end_color=palette.zebra, fill_type="solid")
    subtotal_fill = PatternFill(start_color=palette.subtotal, end_color=palette.subtotal, fill_type="solid")
    border = _build_border()

    # Header row styling
    for col_idx in range(1, max_col + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    # Data row styling
    for row_idx, row in enumerate(tab.rows, start=2):
        is_total_row = bool(row) and any(
            isinstance(cell, str) and cell.strip().lower() in _TOTAL_KEYWORDS
            for cell in row[:3]
        )
        for col_idx in range(1, max_col + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if is_total_row:
                cell.font = Font(bold=True, color=palette.text)
                cell.fill = subtotal_fill
            elif row_idx % 2 == 0:
                cell.fill = zebra_fill
            cell.border = border

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{max_col_letter}{max_row}"

    # Number formats
    for col_idx, header in enumerate(tab.headers, start=1):
        col_values = [row[col_idx - 1] for row in tab.rows if col_idx - 1 < len(row)]
        number_format = _infer_number_format(header, col_values)
        if not number_format:
            continue
        for row_idx in range(2, max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if _is_numeric(cell.value):
                cell.number_format = number_format


def _apply_conditional_formatting(ws: Worksheet, tab: SheetTab, palette: SheetPalette) -> None:
    from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
    from openpyxl.utils import get_column_letter

    if not tab.headers or len(tab.rows) < 2:
        return

    max_row = len(tab.rows) + 1
    numeric_cols = _numeric_columns(tab)
    if not numeric_cols:
        return

    for col_idx in numeric_cols:
        num_count = sum(
            1
            for row in tab.rows
            if col_idx < len(row) and _is_numeric(row[col_idx])
        )
        if num_count < 3:
            continue
        col_letter = get_column_letter(col_idx + 1)
        cell_range = f"{col_letter}2:{col_letter}{max_row}"
        ws.conditional_formatting.add(
            cell_range,
            ColorScaleRule(
                start_type="min",
                start_color=palette.trend_down,
                mid_type="percentile",
                mid_value=50,
                mid_color=palette.zebra,
                end_type="max",
                end_color=palette.trend_up,
            ),
        )

    # Add a data bar on the first numeric column as a compact visual.
    first_col = numeric_cols[0]
    col_letter = get_column_letter(first_col + 1)
    ws.conditional_formatting.add(
        f"{col_letter}2:{col_letter}{max_row}",
        DataBarRule(start_type="min", end_type="max", color=palette.accent),
    )


def _numeric_column_metrics(tab: SheetTab) -> List[Dict[str, Any]]:
    metrics: List[Dict[str, Any]] = []
    for col_idx, header in enumerate(tab.headers):
        values = [
            float(row[col_idx])
            for row in tab.rows
            if col_idx < len(row) and _is_numeric(row[col_idx])
        ]
        if not values:
            continue
        metrics.append(
            {
                "header": header,
                "total": round(sum(values), 2),
                "average": round(statistics.mean(values), 2),
            }
        )
    return metrics


def _add_kpi_tiles(ws: Worksheet, tab: SheetTab, palette: SheetPalette, report: Optional[SheetAnalysisReport]) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill

    if not tab.headers:
        return

    # Keep KPI tiles out of any pre-populated columns (including formula-only columns).
    used_col_count = max(len(tab.headers), int(getattr(ws, "max_column", 0) or 0))
    base_col = used_col_count + 2
    tile_fill = PatternFill(start_color=palette.secondary, end_color=palette.secondary, fill_type="solid")
    label_font = Font(color=palette.header_text, bold=True, size=9)
    value_font = Font(color=palette.text, bold=True, size=12)

    metrics = _numeric_column_metrics(tab)[:2]
    for i, metric in enumerate(metrics):
        row = 1 + (i * 3)
        ws.cell(row=row, column=base_col, value=f"{metric['header']} TOTAL").font = label_font
        ws.cell(row=row, column=base_col).fill = tile_fill
        ws.cell(row=row, column=base_col).alignment = Alignment(horizontal="center")

        value_cell = ws.cell(row=row + 1, column=base_col, value=metric["total"])
        value_cell.font = value_font
        value_cell.number_format = "#,##0.00"
        value_cell.alignment = Alignment(horizontal="center")

        ws.cell(row=row + 2, column=base_col, value=f"Avg {metric['average']}").alignment = Alignment(horizontal="center")

    if not report or not report.trends:
        return

    trend_colors = {
        "up": palette.trend_up,
        "down": palette.trend_down,
        "flat": palette.trend_flat,
    }
    trend_symbols = {
        "up": "UP",
        "down": "DOWN",
        "flat": "FLAT",
    }
    trend_start_row = 8
    for i, trend in enumerate(report.trends[:2]):
        row = trend_start_row + i
        fill = PatternFill(
            start_color=trend_colors.get(trend.direction, palette.trend_flat),
            end_color=trend_colors.get(trend.direction, palette.trend_flat),
            fill_type="solid",
        )
        cell = ws.cell(
            row=row,
            column=base_col,
            value=f"{trend.column}: {trend_symbols.get(trend.direction, 'FLAT')}",
        )
        cell.fill = fill
        cell.font = Font(color=palette.text, bold=True, size=9)


def _chart_spec_from_plan(tab: SheetTab, entry: Dict[str, Any]) -> Optional[Dict[str, Any]]:
    chart_type = str(entry.get("chart_type", "")).strip().lower()
    if chart_type not in {"bar", "line", "pie", "scatter"}:
        return None

    spec: Dict[str, Any] = {
        "chart_type": chart_type,
        "title": entry.get("title") or f"{tab.name} Chart",
    }

    if chart_type == "scatter":
        x_idx = _header_index(tab, str(entry.get("x_column", "")))
        y_idx = _header_index(tab, str(entry.get("y_column", "")))
        if x_idx is None or y_idx is None:
            logger.debug(
                "Chart plan scatter columns not found in '%s' headers: x=%s y=%s",
                tab.name, entry.get("x_column"), entry.get("y_column"),
            )
            return None
        spec["x_col"] = x_idx + 1
        spec["y_col"] = y_idx + 1
        return spec

    category_idx = _header_index(tab, str(entry.get("category_column", "")))
    value_cols: List[int] = []
    raw_value_cols = entry.get("value_columns")
    if isinstance(raw_value_cols, list):
        for value_col in raw_value_cols:
            if not isinstance(value_col, str):
                continue
            idx = _header_index(tab, value_col)
            if idx is not None:
                value_cols.append(idx + 1)

    if category_idx is None or not value_cols:
        logger.debug(
            "Chart plan columns not found in '%s' headers: category=%s values=%s",
            tab.name, entry.get("category_column"), entry.get("value_columns"),
        )
        return None

    spec["category_col"] = category_idx + 1
    spec["value_cols"] = value_cols
    return spec


def _infer_chart_spec(tab: SheetTab) -> Optional[Dict[str, Any]]:
    if len(tab.rows) < 2 or not tab.headers:
        return None

    numeric_cols = _numeric_columns(tab)
    if not numeric_cols:
        return None

    categorical_cols = _categorical_columns(tab)

    # Scatter fallback if purely numeric.
    if len(numeric_cols) >= 2 and not categorical_cols:
        return {
            "chart_type": "scatter",
            "title": f"{tab.name}: {tab.headers[numeric_cols[0]]} vs {tab.headers[numeric_cols[1]]}",
            "x_col": numeric_cols[0] + 1,
            "y_col": numeric_cols[1] + 1,
        }

    cat_idx = categorical_cols[0] if categorical_cols else 0
    val_idx = numeric_cols[0]
    cat_values = [row[cat_idx] for row in tab.rows if cat_idx < len(row)]

    if _looks_temporal(cat_values, tab.headers[cat_idx]):
        chart_type = "line"
    else:
        unique_categories = {
            str(v).strip() for v in cat_values if v is not None and str(v).strip()
        }
        if len(unique_categories) <= 8 and len(unique_categories) >= len(cat_values) * 0.7:
            chart_type = "pie"
        else:
            chart_type = "bar"

    return {
        "chart_type": chart_type,
        "title": f"{tab.name}: {tab.headers[val_idx]}",
        "category_col": cat_idx + 1,
        "value_cols": [val_idx + 1],
    }


def _add_chart(ws: Worksheet, tab: SheetTab, spec: Dict[str, Any], anchor: str) -> bool:
    from openpyxl.chart import BarChart, LineChart, PieChart, Reference, ScatterChart, Series

    if len(tab.rows) < 2:
        return False

    max_row = len(tab.rows) + 1
    chart_type = spec.get("chart_type")

    try:
        if chart_type == "bar":
            chart = BarChart()
            chart.type = "col"
            for col in spec["value_cols"]:
                chart.add_data(
                    Reference(ws, min_col=col, min_row=1, max_row=max_row),
                    titles_from_data=True,
                )
            chart.set_categories(Reference(ws, min_col=spec["category_col"], min_row=2, max_row=max_row))
        elif chart_type == "line":
            chart = LineChart()
            for col in spec["value_cols"]:
                chart.add_data(
                    Reference(ws, min_col=col, min_row=1, max_row=max_row),
                    titles_from_data=True,
                )
            chart.set_categories(Reference(ws, min_col=spec["category_col"], min_row=2, max_row=max_row))
        elif chart_type == "pie":
            chart = PieChart()
            value_col = spec["value_cols"][0]
            chart.add_data(
                Reference(ws, min_col=value_col, min_row=1, max_row=max_row),
                titles_from_data=True,
            )
            chart.set_categories(Reference(ws, min_col=spec["category_col"], min_row=2, max_row=max_row))
        elif chart_type == "scatter":
            chart = ScatterChart()
            x_ref = Reference(ws, min_col=spec["x_col"], min_row=2, max_row=max_row)
            y_ref = Reference(ws, min_col=spec["y_col"], min_row=2, max_row=max_row)
            series = Series(y_ref, xvalues=x_ref, title=spec.get("title", "Series"))
            chart.series.append(series)
        else:
            return False
    except Exception as exc:
        logger.warning("Chart creation failed (%s): %s", spec.get("chart_type"), exc)
        return False

    chart.title = spec.get("title") or "Chart"
    chart.width = 9.0
    chart.height = 5.2
    ws.add_chart(chart, anchor)
    return True


def _collect_plan_entries(content_tree: SheetContentTree) -> List[Dict[str, Any]]:
    metadata = content_tree.metadata if isinstance(content_tree.metadata, dict) else {}
    chart_plan = metadata.get("chart_plan") if metadata else None
    if not isinstance(chart_plan, list):
        return []

    plan_entries: List[Dict[str, Any]] = []
    for entry in chart_plan:
        if isinstance(entry, dict):
            plan_entries.append(entry)
    return plan_entries


def _add_workbook_charts(
    content_tree: SheetContentTree,
    sheet_refs: List[Dict[str, Any]],
    max_charts: int,
) -> int:
    from openpyxl.utils import get_column_letter

    chart_count = 0
    chartable_exists = False
    anchors_by_sheet: Dict[str, int] = {}

    for ref in sheet_refs:
        tab = ref["tab"]
        if _infer_chart_spec(tab) is not None:
            chartable_exists = True
        anchors_by_sheet[ref["sheet_name"]] = max(14, len(tab.rows) + 3)

    plan_entries = _collect_plan_entries(content_tree)
    if plan_entries:
        for entry in plan_entries:
            if chart_count >= max_charts:
                break
            tab_name = str(entry.get("tab_name", "")).strip().lower()
            target = None
            for ref in sheet_refs:
                if ref["tab"].name.strip().lower() == tab_name:
                    target = ref
                    break
            if target is None:
                continue
            spec = _chart_spec_from_plan(target["tab"], entry)
            if not spec:
                continue
            anchor_col = get_column_letter(len(target["tab"].headers) + 3)
            anchor_row = anchors_by_sheet[target["sheet_name"]]
            if _add_chart(target["ws"], target["tab"], spec, f"{anchor_col}{anchor_row}"):
                anchors_by_sheet[target["sheet_name"]] += 16
                chart_count += 1

    if chart_count >= max_charts:
        return chart_count

    prioritized = sorted(
        sheet_refs,
        key=lambda ref: _preferred_tab_score(ref["tab"].name),
        reverse=True,
    )

    for ref in prioritized:
        if chart_count >= max_charts:
            break
        spec = _infer_chart_spec(ref["tab"])
        if not spec:
            continue
        anchor_col = get_column_letter(len(ref["tab"].headers) + 3)
        anchor_row = anchors_by_sheet[ref["sheet_name"]]
        if _add_chart(ref["ws"], ref["tab"], spec, f"{anchor_col}{anchor_row}"):
            anchors_by_sheet[ref["sheet_name"]] += 16
            chart_count += 1

    # Ensure at least one chart in chartable workbooks when profile allows.
    if chartable_exists and chart_count == 0 and max_charts > 0:
        for ref in sheet_refs:
            spec = _infer_chart_spec(ref["tab"])
            if not spec:
                continue
            anchor_col = get_column_letter(len(ref["tab"].headers) + 3)
            anchor_row = anchors_by_sheet[ref["sheet_name"]]
            if _add_chart(ref["ws"], ref["tab"], spec, f"{anchor_col}{anchor_row}"):
                chart_count = 1
                break

    return chart_count


def export_to_xlsx(content_tree: SheetContentTree, output_path: Path) -> None:
    """Export a SheetContentTree to XLSX format.

    Creates one worksheet per tab with headers, data, formulas, visual styling,
    conditional formatting, charts, and lightweight KPI/trend graphics.
    """
    import openpyxl
    from openpyxl.styles import PatternFill
    from openpyxl.utils import get_column_letter

    if not content_tree.tabs:
        raise ValueError("Cannot export empty sheet (no tabs)")

    output_path.parent.mkdir(parents=True, exist_ok=True)

    palette = _select_palette(content_tree)
    visual_profile = _normalize_visual_profile(content_tree.metadata if isinstance(content_tree.metadata, dict) else None)
    max_charts = _PROFILE_MAX_CHARTS.get(visual_profile, 3)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    used_names = set()
    sheet_refs: List[Dict[str, Any]] = []

    for tab in content_tree.tabs:
        name = sanitize_sheet_name(tab.name)
        base_name = name
        counter = 2
        while name in used_names:
            suffix = f" ({counter})"
            name = base_name[: _MAX_SHEET_NAME_LEN - len(suffix)] + suffix
            counter += 1
        used_names.add(name)

        ws = wb.create_sheet(title=name)

        if tab.headers:
            for col_idx, header in enumerate(tab.headers, start=1):
                ws.cell(row=1, column=col_idx, value=header)

        for row_idx, row_data in enumerate(tab.rows, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        for cell_addr, formula in tab.formulas.items():
            ws[cell_addr] = formula

        if tab.column_widths:
            for col_idx, width in enumerate(tab.column_widths, start=1):
                col_letter = get_column_letter(col_idx)
                ws.column_dimensions[col_letter].width = max(width / 7, 8)
        else:
            for col_idx, header in enumerate(tab.headers, start=1):
                col_letter = get_column_letter(col_idx)
                ws.column_dimensions[col_letter].width = max(len(str(header)) + 4, 12)

        # Subtle background color for all used cells keeps tabs cohesive.
        if tab.headers:
            bg_fill = PatternFill(start_color=palette.background, end_color=palette.background, fill_type="solid")
            for row in ws.iter_rows(min_row=1, max_row=max(1, len(tab.rows) + 1), min_col=1, max_col=len(tab.headers)):
                for cell in row:
                    if cell.fill is None or cell.fill.fill_type is None:
                        cell.fill = bg_fill

        _apply_structural_formatting(ws, tab, palette)
        _apply_conditional_formatting(ws, tab, palette)
        _add_kpi_tiles(ws, tab, palette, content_tree.analysis_report)

        sheet_refs.append({"tab": tab, "ws": ws, "sheet_name": name})

    _add_workbook_charts(content_tree, sheet_refs, max_charts=max_charts)

    wb.save(str(output_path))
