"""Tests for core/studio/sheets/exporter_xlsx.py — XLSX export via openpyxl."""

import openpyxl

from core.schemas.studio_schema import SheetContentTree, SheetTab
from core.studio.sheets.exporter_xlsx import _PALETTE_BY_ID, _select_palette, export_to_xlsx


def _make_tree(**kwargs) -> SheetContentTree:
    defaults = {
        "workbook_title": "Test Workbook",
        "tabs": [
            SheetTab(
                id="t1",
                name="Revenue",
                headers=["Month", "Amount"],
                rows=[["Jan", 100], ["Feb", 200]],
                formulas={"C2": "=B2*1.1"},
                column_widths=[120, 80],
            )
        ],
        "metadata": {"visual_profile": "balanced"},
    }
    defaults.update(kwargs)
    return SheetContentTree(**defaults)


def test_xlsx_creates_file(tmp_path):
    path = tmp_path / "output.xlsx"
    export_to_xlsx(_make_tree(), path)
    assert path.exists()


def test_xlsx_creates_parent_dirs(tmp_path):
    path = tmp_path / "sub" / "dir" / "output.xlsx"
    export_to_xlsx(_make_tree(), path)
    assert path.exists()


def test_xlsx_opens_with_openpyxl(tmp_path):
    path = tmp_path / "output.xlsx"
    export_to_xlsx(_make_tree(), path)
    wb = openpyxl.load_workbook(str(path))
    wb.close()


def test_xlsx_sheet_count_matches_tabs(tmp_path):
    tree = _make_tree(
        tabs=[
            SheetTab(id="t1", name="Sheet1", headers=["A"], rows=[[1]]),
            SheetTab(id="t2", name="Sheet2", headers=["B"], rows=[[2]]),
        ]
    )
    path = tmp_path / "output.xlsx"
    export_to_xlsx(tree, path)
    wb = openpyxl.load_workbook(str(path))
    assert len(wb.sheetnames) == 2
    wb.close()


def test_xlsx_sheet_names_match(tmp_path):
    tree = _make_tree(
        tabs=[
            SheetTab(id="t1", name="Revenue", headers=["A"], rows=[[1]]),
            SheetTab(id="t2", name="Costs", headers=["B"], rows=[[2]]),
        ]
    )
    path = tmp_path / "output.xlsx"
    export_to_xlsx(tree, path)
    wb = openpyxl.load_workbook(str(path))
    assert wb.sheetnames == ["Revenue", "Costs"]
    wb.close()


def test_xlsx_header_row_present(tmp_path):
    path = tmp_path / "output.xlsx"
    export_to_xlsx(_make_tree(), path)
    wb = openpyxl.load_workbook(str(path))
    ws = wb.active
    headers = [ws.cell(row=1, column=c).value for c in range(1, 3)]
    assert headers == ["Month", "Amount"]
    wb.close()


def test_xlsx_data_rows_present(tmp_path):
    path = tmp_path / "output.xlsx"
    export_to_xlsx(_make_tree(), path)
    wb = openpyxl.load_workbook(str(path))
    ws = wb.active
    row1 = [ws.cell(row=2, column=c).value for c in range(1, 3)]
    assert row1 == ["Jan", 100]
    wb.close()


def test_xlsx_formulas_written(tmp_path):
    path = tmp_path / "output.xlsx"
    export_to_xlsx(_make_tree(), path)
    wb = openpyxl.load_workbook(str(path))
    ws = wb.active
    assert ws["C2"].value == "=B2*1.1"
    wb.close()


def test_xlsx_kpi_tiles_do_not_overwrite_existing_side_formula(tmp_path):
    path = tmp_path / "output.xlsx"
    tree = _make_tree(
        tabs=[
            SheetTab(
                id="t1",
                name="Revenue",
                headers=["Month", "Amount"],
                rows=[["Jan", 100], ["Feb", 200], ["Mar", 300]],
                formulas={"D2": "=B2*2"},
                column_widths=[120, 80],
            )
        ]
    )
    export_to_xlsx(tree, path)
    wb = openpyxl.load_workbook(str(path))
    ws = wb.active
    assert ws["D2"].value == "=B2*2"
    assert ws["F1"].value == "Amount TOTAL"
    wb.close()


def test_xlsx_column_widths_applied(tmp_path):
    path = tmp_path / "output.xlsx"
    export_to_xlsx(_make_tree(), path)
    wb = openpyxl.load_workbook(str(path))
    ws = wb.active
    # Column A should have a width based on 120/7
    assert ws.column_dimensions["A"].width > 0
    wb.close()


def test_xlsx_frozen_pane(tmp_path):
    path = tmp_path / "output.xlsx"
    export_to_xlsx(_make_tree(), path)
    wb = openpyxl.load_workbook(str(path))
    ws = wb.active
    assert ws.freeze_panes == "A2"
    wb.close()


def test_xlsx_header_style_applied(tmp_path):
    path = tmp_path / "output.xlsx"
    export_to_xlsx(_make_tree(), path)
    wb = openpyxl.load_workbook(str(path))
    ws = wb.active
    header_cell = ws["A1"]
    assert header_cell.font.bold is True
    assert header_cell.fill.fill_type == "solid"
    wb.close()


def test_xlsx_auto_filter_applied(tmp_path):
    path = tmp_path / "output.xlsx"
    export_to_xlsx(_make_tree(), path)
    wb = openpyxl.load_workbook(str(path))
    ws = wb.active
    assert ws.auto_filter.ref is not None
    wb.close()


def test_xlsx_adds_conditional_formatting_for_numeric_columns(tmp_path):
    path = tmp_path / "output.xlsx"
    export_to_xlsx(_make_tree(), path)
    wb = openpyxl.load_workbook(str(path))
    ws = wb.active
    rule_count = sum(len(cf.rules) for cf in ws.conditional_formatting)
    assert rule_count > 0
    wb.close()


def test_xlsx_balanced_profile_adds_chart_for_chartable_data(tmp_path):
    path = tmp_path / "output.xlsx"
    tree = _make_tree(
        tabs=[
            SheetTab(
                id="t1",
                name="Summary",
                headers=["Month", "Revenue"],
                rows=[["Jan", 120], ["Feb", 180], ["Mar", 260], ["Apr", 220]],
                formulas={},
                column_widths=[120, 100],
            )
        ]
    )
    export_to_xlsx(tree, path)
    wb = openpyxl.load_workbook(str(path))
    ws = wb.active
    assert len(getattr(ws, "_charts", [])) >= 1
    wb.close()


def test_xlsx_non_chartable_data_exports_without_failure(tmp_path):
    path = tmp_path / "output.xlsx"
    tree = _make_tree(
        tabs=[
            SheetTab(
                id="t1",
                name="Notes",
                headers=["Topic", "Owner"],
                rows=[["Kickoff", "Alice"], ["Review", "Bob"]],
                formulas={},
                column_widths=[120, 120],
            )
        ]
    )
    export_to_xlsx(tree, path)
    wb = openpyxl.load_workbook(str(path))
    ws = wb.active
    assert ws["A2"].value == "Kickoff"
    assert len(getattr(ws, "_charts", [])) == 0
    wb.close()


def test_xlsx_invalid_chart_plan_falls_back_to_heuristics(tmp_path):
    path = tmp_path / "output.xlsx"
    tree = _make_tree(
        metadata={
            "visual_profile": "balanced",
            "chart_plan": [
                {
                    "tab_name": "Revenue",
                    "chart_type": "line",
                    "category_column": "Missing",
                    "value_columns": ["Also Missing"],
                }
            ],
        },
        tabs=[
            SheetTab(
                id="t1",
                name="Revenue",
                headers=["Month", "Amount"],
                rows=[["Jan", 100], ["Feb", 200], ["Mar", 320]],
                formulas={},
                column_widths=[120, 80],
            )
        ],
    )
    export_to_xlsx(tree, path)
    wb = openpyxl.load_workbook(str(path))
    ws = wb.active
    assert len(getattr(ws, "_charts", [])) >= 1
    wb.close()


def test_xlsx_palette_selection_is_deterministic():
    tree_a = _make_tree(workbook_title="Quarterly Finance Report")
    tree_b = _make_tree(workbook_title="Quarterly Finance Report")
    assert _select_palette(tree_a).id == _select_palette(tree_b).id


def test_xlsx_palette_hint_overrides_hash():
    tree = _make_tree(
        workbook_title="Anything",
        metadata={"palette_hint": "copper-report"},
    )
    assert _select_palette(tree).id == "copper-report"


def test_xlsx_conditional_formatting_uses_palette_colors(tmp_path):
    path = tmp_path / "output.xlsx"
    tree = _make_tree(
        metadata={"visual_profile": "balanced", "palette_hint": "forest-ledger"},
        tabs=[
            SheetTab(
                id="t1",
                name="Data",
                headers=["Month", "Revenue", "Cost"],
                rows=[["Jan", 100, 80], ["Feb", 200, 150], ["Mar", 300, 220]],
                formulas={},
                column_widths=[100, 80, 80],
            )
        ],
    )
    export_to_xlsx(tree, path)
    wb = openpyxl.load_workbook(str(path))
    ws = wb.active

    palette = _PALETTE_BY_ID["forest-ledger"]
    found_palette_color = False
    for cf in ws.conditional_formatting:
        for rule in cf.rules:
            color_scale = getattr(rule, "colorScale", None)
            if color_scale is None:
                continue
            colors = [c.rgb for c in color_scale.color if hasattr(c, "rgb") and c.rgb]
            for c in colors:
                hex_part = c[-6:] if len(c) > 6 else c
                if hex_part == palette.trend_up or hex_part == palette.trend_down:
                    found_palette_color = True

    assert found_palette_color, "ColorScale should use palette trend colors, not hardcoded hex"
    wb.close()


def test_xlsx_chart_anchor_below_data_rows(tmp_path):
    path = tmp_path / "output.xlsx"
    rows = [[f"Item {i}", i * 10] for i in range(1, 21)]
    tree = _make_tree(
        tabs=[
            SheetTab(
                id="t1",
                name="Large",
                headers=["Name", "Value"],
                rows=rows,
                formulas={},
                column_widths=[120, 80],
            )
        ],
    )
    export_to_xlsx(tree, path)
    wb = openpyxl.load_workbook(str(path))
    ws = wb.active
    charts = getattr(ws, "_charts", [])
    assert len(charts) >= 1, "Expected at least one chart for numeric data"

    anchor = charts[0].anchor
    # openpyxl uses OneCellAnchor with _from.row (0-based row index)
    anchor_row_0based = anchor._from.row
    anchor_row = anchor_row_0based + 1  # convert to 1-based
    data_end_row = len(rows) + 1  # +1 for header
    assert anchor_row > data_end_row, (
        f"Chart anchor row {anchor_row} should be below data end row {data_end_row}"
    )
    wb.close()


def test_xlsx_pie_chart_skipped_for_duplicate_categories(tmp_path):
    """When category column has many duplicates, should use bar instead of pie."""
    path = tmp_path / "output.xlsx"
    tree = _make_tree(
        tabs=[
            SheetTab(
                id="t1",
                name="Data",
                headers=["Category", "Metric", "Value"],
                rows=[
                    ["General", "Metric A", 100],
                    ["General", "Metric B", 200],
                    ["General", "Metric C", 150],
                    ["Pricing", "Metric D", 300],
                    ["Pricing", "Metric E", 250],
                    ["Cost", "Metric F", 80],
                ],
                formulas={},
                column_widths=[120, 120, 80],
            )
        ],
    )
    export_to_xlsx(tree, path)
    wb = openpyxl.load_workbook(str(path))
    ws = wb.active
    charts = getattr(ws, "_charts", [])
    # Should have a chart, but NOT a pie chart (categories have duplicates)
    for chart in charts:
        assert not isinstance(chart, openpyxl.chart.PieChart), \
            "Pie chart should not be used when categories have many duplicates"
    wb.close()


def test_xlsx_mrr_column_gets_currency_format(tmp_path):
    path = tmp_path / "output.xlsx"
    tree = _make_tree(
        tabs=[
            SheetTab(
                id="t1",
                name="Revenue",
                headers=["Month", "Total MRR", "Churn Rate"],
                rows=[["Jan", 7900.123, 0.02], ["Feb", 10946.456, 0.03]],
                formulas={},
                column_widths=[100, 100, 100],
            )
        ]
    )
    export_to_xlsx(tree, path)
    wb = openpyxl.load_workbook(str(path))
    ws = wb.active
    # "Total MRR" should get currency format (contains "mrr")
    mrr_cell = ws.cell(row=2, column=2)
    assert "$" in mrr_cell.number_format or "#,##0" in mrr_cell.number_format, \
        f"MRR column should have currency/number format, got: {mrr_cell.number_format}"
    # "Churn Rate" should get percent format (contains "churn" or "rate")
    churn_cell = ws.cell(row=2, column=3)
    assert "%" in churn_cell.number_format, \
        f"Churn Rate column should have percent format, got: {churn_cell.number_format}"
    wb.close()
